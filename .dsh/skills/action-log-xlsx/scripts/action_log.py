#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
行动表 Excel 记录脚本（空想圣杯 · 圣杯玩家指令总控.xlsx 格式）

把 `.从者行动 ...` / `.御主行动 ...` 解析后的行动写进行动表的对应格子。

用法：
  python action_log.py init   --file <xlsx> [--from <模板xlsx>] [--force] [--clear-roster]
  python action_log.py status --file <xlsx>
  python action_log.py record --file <xlsx> --role SERVANT|MASTER --class 弓 --text "机动-灵脉-C"
                              [--day N --period DAY|NIGHT] [--overwrite] [--state <json>] [--dry-run]
  python action_log.py set-day --file <xlsx> --day N --period DAY|NIGHT [--state <json>]
  python action_log.py handoff --file <xlsx> [--out <md路径>] [--title <标题>] [--note <附加说明>]

说明：
- 格子定位：行动记录区位于 Sheet1 第 35~91 行，列 F..L 对应 弓/杀/骑/枪/剑/术/狂；
  第 35 行为「跳伞」，之后每天两行（昼=从者/下一行=御主，夜=从者/下一行=御主）。
- day: 1~14；period: JUMP(跳伞, day=0) / DAY / NIGHT；role: SERVANT / MASTER。
- 默认「追加合并」：格子已有内容时用「，」连接；--overwrite 则整体覆盖。
- --state 指向状态 json（记录当前进行到第几天/昼或夜），缺省为 xlsx 同目录下 <文件名>.state.json。
"""

import argparse
import datetime
import json
import os
import shutil
import sys

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

import openpyxl

SHEET_NAME = 'Sheet1'

# 行动记录区的职阶列（与表头 F34..L34 顺序一致）
CLASS_COLUMNS = {'弓': 'F', '杀': 'G', '骑': 'H', '枪': 'I', '剑': 'J', '术': 'K', '狂': 'L'}

# 基本信息区的职阶列（E4 表头，含「仇」）
ROSTER_COLUMNS = {'弓': 'F', '杀': 'G', '骑': 'H', '枪': 'I', '剑': 'J', '术': 'K', '狂': 'L', '仇': 'M'}

JUMP_ROW = 35
# 每天 4 行：昼从者/昼御主/夜从者/夜御主
#   第1天: 36/37/38/39；第2天: 40/41/42/43；第3天: 44/45/46/47 …
DAY_SERVANT_ROW = lambda day: 32 + 4 * day
DAY_MASTER_ROW = lambda day: 33 + 4 * day
NIGHT_SERVANT_ROW = lambda day: 34 + 4 * day
NIGHT_MASTER_ROW = lambda day: 35 + 4 * day

def _find_project_dir():
    """从脚本目录向上查找包含 1.15 目录的项目根"""
    d = os.path.dirname(os.path.abspath(__file__))
    for _ in range(8):
        if os.path.isdir(os.path.join(d, '1.15')):
            return d
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))

DEFAULT_TEMPLATE = os.path.join(_find_project_dir(), '1.15', '圣杯玩家指令总控.xlsx')


def state_path_for(xlsx_path):
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    return os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), base + '.state.json')


def load_state(xlsx_path, state_arg):
    path = state_arg or state_path_for(xlsx_path)
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_state(xlsx_path, state_arg, data):
    path = state_arg or state_path_for(xlsx_path)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def resolve_row(day, period, role):
    if period == 'JUMP':
        return JUMP_ROW
    if period == 'DAY':
        return DAY_SERVANT_ROW(day) if role == 'SERVANT' else DAY_MASTER_ROW(day)
    if period == 'NIGHT':
        return NIGHT_SERVANT_ROW(day) if role == 'SERVANT' else NIGHT_MASTER_ROW(day)
    raise ValueError(f'period 非法: {period}')


def cmd_init(args):
    if os.path.exists(args.file) and not args.force:
        print(f'[init] 文件已存在: {args.file}（跳过；如需重建请加 --force）')
        return 0
    src = args.from_path or DEFAULT_TEMPLATE
    if not os.path.exists(src):
        print(f'[init] 模板不存在: {src}', file=sys.stderr)
        return 1
    os.makedirs(os.path.dirname(os.path.abspath(args.file)), exist_ok=True)
    shutil.copy2(src, args.file)
    wb = openpyxl.load_workbook(args.file)
    ws = wb[SHEET_NAME]
    # 清空行动记录区（跳伞 + 第1~14天 × 昼夜，列 F..L）
    cleared = 0
    for row in range(JUMP_ROW, 92):
        for col in 'FGHIJKL':
            if ws[f'{col}{row}'].value not in (None, ''):
                ws[f'{col}{row}'].value = None
                cleared += 1
    if args.clear_roster:
        for row in range(4, 30):
            for col in 'FGHIJKLM':
                if ws[f'{col}{row}'].value not in (None, ''):
                    ws[f'{col}{row}'].value = None
    wb.save(args.file)
    print(f'[init] 已从模板创建: {args.file}')
    print(f'[init] 已清空行动记录格 {cleared} 个' + ('，并清空基本信息名单' if args.clear_roster else '（基本信息名单保留模板内容，需要清空请加 --clear-roster）'))
    return 0


def cmd_status(args):
    if not os.path.exists(args.file):
        print(f'[status] 文件不存在: {args.file}（先运行 init）', file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[SHEET_NAME]
    print(f'== 行动表: {args.file} ==')
    # 基本信息
    try:
        for label_row, label in ((4, '阶职'), (5, '从者真名'), (7, '御主真名'), (11, '从者简称'), (12, '御主简称'), (28, '从者玩家'), (29, '御主玩家')):
            parts = []
            for cls, col in ROSTER_COLUMNS.items():
                v = ws[f'{col}{label_row}'].value
                if v is not None and str(v).strip():
                    parts.append(f'{cls}={v}')
            if parts:
                print(f'  基本信息·{label}: ' + '，'.join(parts))
    except Exception as e:
        print(f'  （基本信息读取失败: {e}）')
    # 当前状态
    state = load_state(args.file, args.state)
    if state.get('day') is not None:
        print(f'  状态文件: 当前进行到 第{state["day"]}天 {state.get("period", "")}')
    # 行动记录概览
    print('  行动记录区（非空格子）:')
    empty = True
    for row in range(JUMP_ROW, 92):
        if row == JUMP_ROW:
            label = '跳伞'
        else:
            idx = (row - 36) // 2  # 0=第1天昼(36,37), 1=第1天夜(38,39), 2=第2天昼(40,41)...
            day = idx // 2 + 1
            period_name = '昼' if idx % 2 == 0 else '夜'
            role_name = '从者' if row % 2 == 0 else '御主'
            label = f'第{day}天{period_name}·{role_name}'
        cells = []
        for cls, col in CLASS_COLUMNS.items():
            v = ws[f'{col}{row}'].value
            if v is not None and str(v).strip():
                cells.append(f'{cls}:{v}')
        if cells:
            empty = False
            print(f'    {label}: ' + ' | '.join(cells))
    if empty:
        print('    （无记录）')
    wb.close()
    return 0


def cmd_record(args):
    if not os.path.exists(args.file):
        print(f'[record] 文件不存在: {args.file}（先运行 init）', file=sys.stderr)
        return 1
    cls = args.class_name
    if cls not in CLASS_COLUMNS:
        print(f'[record] 未知职阶「{cls}」，可用: {", ".join(CLASS_COLUMNS)}', file=sys.stderr)
        return 1
    role = args.role.upper()
    if role not in ('SERVANT', 'MASTER'):
        print(f'[record] role 非法: {args.role}', file=sys.stderr)
        return 1
    text = (args.text or '').strip()
    if not text:
        print('[record] text 不能为空', file=sys.stderr)
        return 1

    # 天/时段：优先命令行，其次状态文件，缺省第1天昼
    state = load_state(args.file, args.state)
    day = args.day if args.day is not None else state.get('day', 1)
    period = args.period or state.get('period', 'DAY')
    if period != 'JUMP' and not (1 <= day <= 14):
        print(f'[record] day 非法: {day}（应为 1~14；跳伞请用 --period JUMP --day 0）', file=sys.stderr)
        return 1

    col = CLASS_COLUMNS[cls]
    row = resolve_row(day, period, role)
    period_label = '跳伞' if period == 'JUMP' else f'第{day}天{"昼" if period == "DAY" else "夜"}'

    wb = openpyxl.load_workbook(args.file)
    ws = wb[SHEET_NAME]
    cell = ws[f'{col}{row}']
    old = str(cell.value).strip() if cell.value is not None else ''

    if old and not args.overwrite:
        # 追加合并，去重
        if text in old:
            print(f'[record] 内容已存在，未改动: {col}{row} = {old}')
            wb.close()
            return 0
        new = old + '，' + text
        action = '合并追加'
    else:
        new = text
        action = '覆盖写入' if old else '写入'

    if args.dry_run:
        print(f'[dry-run] {action} {col}{row}（{period_label} · {role} · {cls}）')
        print(f'         内容: {new}')
        wb.close()
        return 0

    cell.value = new
    wb.save(args.file)
    wb.close()

    # 更新状态文件：记录写入后的 天/时段（跳伞后推进到第1天昼）
    new_state = dict(state)
    if period == 'JUMP':
        new_state.update({'day': 1, 'period': 'DAY'})
    else:
        new_state.update({'day': day, 'period': period})
    sp = save_state(args.file, args.state, new_state)

    print(f'[record] {action}成功: {col}{row}（{period_label} · {role} · {cls}）')
    print(f'         内容: {new}')
    print(f'         状态已更新: {sp}（第{new_state["day"]}天 {new_state["period"]}）')
    return 0


def cmd_set_day(args):
    if args.period == 'JUMP':
        day, period = 0, 'JUMP'
    else:
        day, period = args.day, args.period
    if period != 'JUMP' and not (1 <= day <= 14):
        print(f'[set-day] day 非法: {day}', file=sys.stderr)
        return 1
    state = load_state(args.file, args.state)
    state.update({'day': day, 'period': period})
    sp = save_state(args.file, args.state, state)
    print(f'[set-day] 已设为 第{day}天 {period}（状态文件: {sp}）')
    return 0


def cmd_advance(args):
    """按顺序推进时段：跳伞 → 第1天昼 → 第1天夜 → 第2天昼 → … → 第14天夜"""
    state = load_state(args.file, args.state)
    day = state.get('day')
    period = state.get('period')
    if day is None or period is None:
        n_day, n_period = 1, 'DAY'
        cur = '（未设置，从跳伞后开始）'
    elif period == 'JUMP':
        n_day, n_period = 1, 'DAY'
        cur = '跳伞'
    elif period == 'DAY':
        n_day, n_period = day, 'NIGHT'
        cur = f'第{day}天昼'
    elif period == 'NIGHT':
        if day >= 14:
            print('[advance] 已是最后时段：第14天夜，无法继续推进', file=sys.stderr)
            return 1
        n_day, n_period = day + 1, 'DAY'
        cur = f'第{day}天夜'
    else:
        print(f'[advance] 状态文件时段非法: {period}', file=sys.stderr)
        return 1
    state.update({'day': n_day, 'period': n_period})
    sp = save_state(args.file, args.state, state)
    print(f'[advance] 已从 {cur} 推进到 第{n_day}天 {n_period}（状态文件: {sp}）')
    return 0


def cmd_chatlog(args):
    """把机器人落盘的群聊 JSONL 转成可读聊天记录（AI 总结填表用）"""
    import json as _json
    if not os.path.exists(args.file):
        print(f'[错误] 群聊记录文件不存在: {args.file}', file=sys.stderr)
        return 1
    recs = []
    with open(args.file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                recs.append(_json.loads(line))
            except Exception:
                continue
    if args.guild:
        recs = [r for r in recs if str(r.get('guildId', '')) == str(args.guild)]
    shown = recs[-args.lines:] if args.lines else recs
    print(f'== 群聊记录: {os.path.basename(args.file)}（共 {len(recs)} 条，显示最近 {len(shown)} 条）==')
    for r in shown:
        try:
            ts = datetime.datetime.fromisoformat(str(r.get('ts', '')).replace('Z', '+00:00')).astimezone()
            t = ts.strftime('%m-%d %H:%M')
        except Exception:
            t = str(r.get('ts', ''))[:16]
        nick = str(r.get('nickname') or r.get('userId') or '?')
        text = str(r.get('text', ''))
        print(f'{t} [{nick}] {text}')
    if not shown:
        print('（无记录）')
    return 0


def _period_label(day, period):
    if period == 'JUMP':
        return '跳伞'
    return f'第{day}天{"昼" if period == "DAY" else "夜"}'


def cmd_handoff(args):
    """生成/更新交接文档（markdown），让下一个会话快速接上进度"""
    if not os.path.exists(args.file):
        print(f'[handoff] 文件不存在: {args.file}', file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[SHEET_NAME]

    title = args.title or os.path.splitext(os.path.basename(args.file))[0]
    out = args.out or os.path.join(os.path.dirname(os.path.abspath(args.file)), title + '交接.md')

    # 收集行动记录（跳伞 + 第1~14天昼夜，从者/御主合并显示）
    rows_data = []
    for row in (JUMP_ROW,):
        cells = [str(ws[f'{col}{row}'].value).strip() if ws[f'{col}{row}'].value is not None else ''
                 for col in 'FGHIJKL']
        rows_data.append(('跳伞', cells))
    for day in range(1, 15):
        for pname, srow, mrow in (('昼', DAY_SERVANT_ROW(day), DAY_MASTER_ROW(day)),
                                  ('夜', NIGHT_SERVANT_ROW(day), NIGHT_MASTER_ROW(day))):
            cells = []
            for col in 'FGHIJKL':
                s = ws[f'{col}{srow}'].value
                m = ws[f'{col}{mrow}'].value
                s = str(s).strip() if s is not None else ''
                m = str(m).strip() if m is not None else ''
                if s and m:
                    cells.append(f'从:{s}；御:{m}')
                elif s:
                    cells.append(f'从:{s}')
                elif m:
                    cells.append(f'御:{m}')
                else:
                    cells.append('')
            rows_data.append((f'第{day}天{pname}', cells))

    filled_count = sum(1 for _, cells in rows_data for c in cells if c)
    state = load_state(args.file, args.state)
    cur = f'第{state.get("day", 1)}天 {state.get("period", "DAY")}' if state.get('day') is not None else '未设置（默认从第1天昼开始）'

    lines = []
    lines.append(f'# {title}交接')
    lines.append('')
    lines.append(f'> 更新时间：{datetime.datetime.now():%Y-%m-%d %H:%M}')
    lines.append(f'> 行动表：`{os.path.abspath(args.file)}`')
    lines.append('')
    lines.append('## 当前进度')
    lines.append(f'- 当前进行到：{cur}')
    lines.append(f'- 已记录：{filled_count} 个格子')
    lines.append('')
    lines.append('## 行动记录总览')
    lines.append('| 时段 | 弓 | 杀 | 骑 | 枪 | 剑 | 术 | 狂 |')
    lines.append('|---|---|---|---|---|---|---|---|')
    for label, cells in rows_data:
        if any(cells):
            lines.append('| ' + label + ' | ' + ' | '.join(cells) + ' |')
    if filled_count == 0:
        lines.append('| （暂无记录） | | | | | | | |')
    lines.append('')
    lines.append('## 名单')
    for cls in list(CLASS_COLUMNS) + ['仇']:
        col = ROSTER_COLUMNS[cls]
        parts = []
        for label_row, prefix in ((5, '从者'), (7, '御主'), (28, '玩家')):
            v = ws[f'{col}{label_row}'].value
            if v is not None and str(v).strip():
                parts.append(f'{prefix}:{str(v).strip()}')
        if parts:
            lines.append(f'- {cls}：' + '，'.join(parts))
    lines.append('')
    lines.append('## 下次继续')
    lines.append('- 直接说 `.从者行动 <内容>` / `.御主行动 <内容>`（如 `机动-灵脉-C`），AI 会用 action-log-xlsx 技能写入并更新本文件。')
    lines.append('- 也可手动查看最新状态：')
    lines.append(f'  `python "{os.path.abspath(__file__)}" status --file "{os.path.abspath(args.file)}"`')
    lines.append('- 行动内容原样写入；格子已有内容自动用「，」合并，重复内容跳过；`--overwrite` 可覆盖。')
    if args.note:
        lines.append('')
        lines.append('## 备注')
        lines.append(args.note)
    lines.append('')
    lines.append('## 相关文件')
    lines.append(f'- 行动表：`{os.path.abspath(args.file)}`')
    lines.append(f'- 状态文件：`{state_path_for(args.file)}`（记录当前天/时段）')
    lines.append(f'- 技能：`.dsh/skills/action-log-xlsx/` 与 `.claude/skills/action-log-xlsx/`')
    lines.append('')

    with open(out, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    wb.close()
    print(f'[handoff] 已更新交接文档: {out}')
    print(f'[handoff] 当前进度: {cur}，已记录 {filled_count} 个格子')
    return 0


def main():
    parser = argparse.ArgumentParser(description='行动表 Excel 记录脚本')
    sub = parser.add_subparsers(dest='command', required=True)

    p_init = sub.add_parser('init')
    p_init.add_argument('--file', required=True, help='目标 xlsx 路径')
    p_init.add_argument('--from', dest='from_path', default=None, help='模板 xlsx（缺省为项目 1.15/圣杯玩家指令总控.xlsx）')
    p_init.add_argument('--force', action='store_true', help='已存在时覆盖重建')
    p_init.add_argument('--clear-roster', action='store_true', help='同时清空基本信息名单')

    p_status = sub.add_parser('status')
    p_status.add_argument('--file', required=True)
    p_status.add_argument('--state', default=None)

    p_record = sub.add_parser('record')
    p_record.add_argument('--file', required=True)
    p_record.add_argument('--role', required=True, help='SERVANT / MASTER')
    p_record.add_argument('--class', dest='class_name', required=True, help='职阶：弓/杀/骑/枪/剑/术/狂')
    p_record.add_argument('--text', required=True, help='行动内容，如 "机动-灵脉-C"')
    p_record.add_argument('--day', type=int, default=None, help='第几天 1~14（缺省用状态文件）')
    p_record.add_argument('--period', default=None, help='DAY / NIGHT / JUMP（缺省用状态文件）')
    p_record.add_argument('--overwrite', action='store_true', help='覆盖整格而不是合并追加')
    p_record.add_argument('--state', default=None)
    p_record.add_argument('--dry-run', action='store_true', help='只预览不写入')

    p_day = sub.add_parser('set-day')
    p_day.add_argument('--file', required=True)
    p_day.add_argument('--day', type=int, required=True)
    p_day.add_argument('--period', required=True, help='DAY / NIGHT / JUMP')
    p_day.add_argument('--state', default=None)

    p_adv = sub.add_parser('advance')
    p_adv.add_argument('--file', required=True)
    p_adv.add_argument('--state', default=None)

    p_chat = sub.add_parser('chatlog')
    p_chat.add_argument('--file', required=True, help='群聊记录 JSONL 路径')
    p_chat.add_argument('--lines', type=int, default=50, help='显示最近 N 条（默认50）')
    p_chat.add_argument('--guild', default=None, help='按群 ID 过滤')

    p_handoff = sub.add_parser('handoff')
    p_handoff.add_argument('--file', required=True)
    p_handoff.add_argument('--out', default=None, help='输出 md 路径（缺省为 xlsx 同目录 <标题>交接.md）')
    p_handoff.add_argument('--title', default=None, help='文档标题（缺省为文件名去掉扩展名）')
    p_handoff.add_argument('--note', default=None, help='附加备注内容')
    p_handoff.add_argument('--state', default=None)

    args = parser.parse_args()
    if args.command == 'init':
        return cmd_init(args)
    if args.command == 'status':
        return cmd_status(args)
    if args.command == 'record':
        return cmd_record(args)
    if args.command == 'set-day':
        return cmd_set_day(args)
    if args.command == 'advance':
        return cmd_advance(args)
    if args.command == 'chatlog':
        return cmd_chatlog(args)
    if args.command == 'handoff':
        return cmd_handoff(args)
    return 1


if __name__ == '__main__':
    sys.exit(main())
