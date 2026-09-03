#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
战斗表 Excel 辅助脚本（空想圣杯 · 战斗表.xlsx）

核心思路：
- 双方战斗表 的属性格是公式，从「戰鬥計算」名单表(M列=角色名, N~T=七项属性) INDEX/MATCH 自动取值；
  所以「自动填属性」= 把角色属性写进名单 + 把名字填进参战位，战斗表会自动算出总值/辅助减半。
- 技能/宝具的判定顺序信息在角色卡里（发动时机：常驻/随时/战斗开始时/初始/主要/最终工序）。

用法：
  python battle_sheet.py status     --file <战斗表.xlsx> [--sheet 双方战斗表]
  python battle_sheet.py roster     --file <战斗表.xlsx>
  python battle_sheet.py fill-stats --file <战斗表.xlsx> --card <角色卡.xlsx> --name <名单角色名>
                                    [--side 蓝|黄 --pos 主力|辅助1|辅助2|辅助3|辅助4] [--no-slot] [--dry-run]
  python battle_sheet.py skills     --card <角色卡.xlsx>
  python battle_sheet.py verdict    --file <判定记录md> --append <文本>
  python battle_sheet.py handoff    --file <战斗表.xlsx> [--out <md>] [--note <文本>] [--verdict <判定记录md>]
"""

import argparse
import datetime
import json
import os
import sys

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

import openpyxl

# ============ 常量：双方战斗表 ============
SHEET_BATTLE = '双方战斗表'
SHEET_CALC = '戰鬥計算'

# 参战位：蓝=B5(主力) C5~F5(辅助)；黄=H5(主力) I5~L5(辅助)
BLUE_SLOTS = {'主力': 'B5', '辅助1': 'C5', '辅助2': 'D5', '辅助3': 'E5', '辅助4': 'F5'}
YELLOW_SLOTS = {'主力': 'H5', '辅助1': 'I5', '辅助2': 'J5', '辅助3': 'K5', '辅助4': 'L5'}

# 属性行：等级7 筋力8 耐力9 敏捷10 魔力11 幸运12 宝具13
STAT_ROWS = {'等级': 7, '筋力': 8, '耐力': 9, '敏捷': 10, '魔力': 11, '幸运': 12, '宝具': 13}

# 名单表（戰鬥計算）：M=角色名, N=等级, O=筋力, P=耐力, Q=敏捷, R=魔力, S=幸运, T=宝具
ROSTER_COLS = {'等级': 'N', '筋力': 'O', '耐力': 'P', '敏捷': 'Q', '魔力': 'R', '幸运': 'S', '宝具': 'T'}
ROSTER_MAX_ROW = 159

# 角色卡属性：合计行（扫描 B 列='合计'），C=等级 D=筋力 E=耐久 F=敏捷 G=魔力 H=幸运 I=宝具
CARD_STAT_COLS = {'等级': 'C', '筋力': 'D', '耐力': 'E', '敏捷': 'F', '魔力': 'G', '幸运': 'H', '宝具': 'I'}


def load_ws(path, sheet):
    if not os.path.exists(path):
        print(f'[错误] 文件不存在: {path}', file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        print(f'[错误] sheet 不存在: {sheet}（可用: {wb.sheetnames}）', file=sys.stderr)
        wb.close()
        sys.exit(1)
    return wb, wb[sheet]


def cell(ws, ref):
    v = ws[ref].value
    return '' if v is None else str(v).strip()


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def row_label(ref):
    """如 B5 → 蓝·主力"""
    col, row = ref[0], int(ref[1:])
    for name, r in BLUE_SLOTS.items():
        if r == ref:
            return f'蓝·{name}'
    for name, r in YELLOW_SLOTS.items():
        if r == ref:
            return f'黄·{name}'
    return ref


# ============ status ============
def cmd_status(args):
    wb, ws = load_ws(args.file, args.sheet)
    print(f'== 战斗表状态: {args.file} [{args.sheet}] ==')

    # 1) 参战人选
    blue = [cell(ws, BLUE_SLOTS[k]) for k in ('主力', '辅助1', '辅助2', '辅助3', '辅助4')]
    yellow = [cell(ws, YELLOW_SLOTS[k]) for k in ('主力', '辅助1', '辅助2', '辅助3', '辅助4')]
    print('【参战人选】')
    print(f'  蓝方: 主力={blue[0]} 辅助={[b for b in blue[1:] if b and b != "无"] or "无"}')
    print(f'  黄方: 主力={yellow[0]} 辅助={[y for y in yellow[1:] if y and y != "无"] or "无"}')

    # 2) 属性区
    print('【属性与总值】（主力+辅助减半合计）')
    for stat, r in STAT_ROWS.items():
        bm, ys = [], []
        for col in 'BCDEF':
            v = cell(ws, f'{col}{r}')
            if v:
                bm.append(v)
        for col in 'HIJKL':
            v = cell(ws, f'{col}{r}')
            if v:
                ys.append(v)
        print(f'  {stat}: 蓝[{",".join(bm)}] 总值={cell(ws, f"G{r}")} | 黄[{",".join(ys)}] 总值={cell(ws, f"M{r}")}')

    # 3) 魔力不足
    print('【魔力不足】')
    print(f'  蓝: {[cell(ws, f"{c}6") for c in "BCDEF"]} | 黄: {[cell(ws, f"{c}6") for c in "HIJKL"]}')
    print('  惩罚档位: ' + '，'.join(f'{cell(ws, f"N{r}")}={cell(ws, f"O{r}")}' for r in range(5, 14) if cell(ws, f"N{r}")))

    # 4) 战前胜补/胜惩
    print('【战前修正】')
    print(f'  胜补: 蓝={cell(ws, "G14")}（明细{[cell(ws, f"{c}14") for c in "BCDEF"]}） | 黄={cell(ws, "M14")}（明细{[cell(ws, f"{c}14") for c in "HIJKL"]}）')
    print(f'  胜惩: 蓝={cell(ws, "G15")}（明细{[cell(ws, f"{c}15") for c in "BCDEF"]}） | 黄={cell(ws, "M15")}（明细{[cell(ws, f"{c}15") for c in "HIJKL"]}）')

    # 5) 当前魔力区（Q~AA）
    print('【当前魔力/消耗区】')
    print(f'  当前魔力: 蓝主力={cell(ws, "R13")} 蓝辅助={[cell(ws, f"{c}13") for c in ("S","T","U","V")]} | 黄主力={cell(ws, "W13")} 黄辅助={[cell(ws, f"{c}13") for c in ("X","Y","Z","AA")]}')
    print(f'  属性扣除={cell(ws, "R14")} 单独扣除={cell(ws, "R15")} | 保底={cell(ws, "R16")} 保穿={cell(ws, "R17")} | 最终: 蓝={cell(ws, "R18")} 黄={cell(ws, "W18")}')

    # 6) 组别A/B 属性区
    print('【组别属性区】')
    for r in range(20, 30):
        label = cell(ws, f'A{r}')
        if label and label != '组别':
            print(f'  {label}: 等级={cell(ws, f"B{r}")} 筋力={cell(ws, f"C{r}")} 耐力={cell(ws, f"D{r}")} 敏捷={cell(ws, f"E{r}")} 魔力={cell(ws, f"F{r}")} 幸运={cell(ws, f"G{r}")} 宝具={cell(ws, f"H{r}")} | 宝外补惩={cell(ws, f"I{r}")} 全属补惩={cell(ws, f"J{r}")} 上三补惩={cell(ws, f"K{r}")}')

    # 7) 属性结算对应胜率
    print('【属性结算→基础胜率】')
    for r in range(45, 61):
        m, n = cell(ws, f'M{r}'), cell(ws, f'N{r}')
        if m:
            print(f'  {m} → {n}%')

    # 8) 最终工序胜率链
    print('【最终工序】')
    headers = {c: cell(ws, f'{c}85') for c in 'BCDEFGHIJKLMN'}
    for c in 'BCDEFGHIJKLMN':
        if headers[c]:
            print(f'  {headers[c]}: 蓝={cell(ws, f"{c}86")} 黄={cell(ws, f"{c}87")}')
    print(f'  工序开关: D84={cell(ws, "D84")} E84={cell(ws, "E84")}')

    # 9) 名单核对
    print('【名单核对（戰鬥計算）】')
    _, cws = load_ws(args.file, SHEET_CALC)
    roster = {}
    for r in range(2, ROSTER_MAX_ROW + 1):
        name = cell(cws, f'M{r}')
        if name:
            roster[name] = {k: cell(cws, f'{v}{r}') for k, v in ROSTER_COLS.items()}
    participants = [x for x in blue + yellow if x and x != '无']
    print(f'  名单共 {len(roster)} 人')
    for p in participants:
        if p in roster:
            st = roster[p]
            print(f'  ✓ {p}: ' + ' '.join(f'{k}={st[k]}' for k in ROSTER_COLS))
        else:
            print(f'  ✗ {p}: 不在名单中（需 fill-stats 或手动加入戰鬥計算）')
    wb.close()
    return 0


# ============ roster ============
def cmd_roster(args):
    _, cws = load_ws(args.file, SHEET_CALC)
    print(f'== 戰鬥計算 名单 ==')
    for r in range(2, ROSTER_MAX_ROW + 1):
        name = cell(cws, f'M{r}')
        if name:
            stats = ' '.join(f'{k}={cell(cws, f"{v}{r}")}' for k, v in ROSTER_COLS.items())
            print(f'  M{r}: {name} | {stats}')
    return 0


# ============ fill-stats ============
def read_card_stats(card_path):
    if not os.path.exists(card_path):
        print(f'[错误] 角色卡不存在: {card_path}', file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(card_path, read_only=True, data_only=True)
    sheet = None
    for cand in ('角色卡 ', '角色卡'):
        if cand in wb.sheetnames:
            sheet = wb[cand]
            break
    if sheet is None:
        print(f'[错误] 角色卡中没有「角色卡」sheet（可用: {wb.sheetnames}）', file=sys.stderr)
        wb.close()
        sys.exit(1)
    # 找到「合计」行
    stats = None
    for r in range(1, 40):
        if str(sheet[f'B{r}'].value or '').strip() == '合计':
            stats = {}
            for k, col in CARD_STAT_COLS.items():
                v = sheet[f'{col}{r}'].value
                stats[k] = '' if v is None else str(v).strip()
            break
    if stats is None:
        print('[错误] 角色卡中未找到「合计」行（属性区在第14~15行附近）', file=sys.stderr)
        wb.close()
        sys.exit(1)
    wb.close()
    return stats


def find_or_append_roster_row(cws, name, write=True):
    for r in range(2, ROSTER_MAX_ROW + 1):
        if cell(cws, f'M{r}') == name:
            return r, False
    for r in range(2, ROSTER_MAX_ROW + 1):
        if not cell(cws, f'M{r}'):
            if write:
                cws[f'M{r}'] = name
            return r, True
    print(f'[错误] 名单已满（{ROSTER_MAX_ROW} 行），请手动清理', file=sys.stderr)
    sys.exit(1)


def cmd_fill_stats(args):
    stats = read_card_stats(args.card)
    if not os.path.exists(args.file):
        print(f'[错误] 文件不存在: {args.file}', file=sys.stderr)
        return 1
    # 必须用 data_only=False 加载并保存，否则公式会被缓存值覆盖
    wb = openpyxl.load_workbook(args.file)
    ws = wb[SHEET_BATTLE]
    cws = wb[SHEET_CALC]
    row, created = find_or_append_roster_row(cws, args.name, write=not args.dry_run)
    print(f'[fill-stats] 角色卡 → 名单 M{row}（{"新增" if created else "更新"}）: {args.name}' + (' [dry-run]' if args.dry_run else ''))
    for k, col in ROSTER_COLS.items():
        print(f'    {k}={stats.get(k, "")}')
    if not args.dry_run:
        for k, col in ROSTER_COLS.items():
            v = stats.get(k, '')
            if v != '':
                cws[f'{col}{row}'] = float(v) if str(v).replace('.', '', 1).isdigit() else v
    # 填参战位
    if not args.no_slot:
        slots = BLUE_SLOTS if args.side == '蓝' else YELLOW_SLOTS
        if args.pos in slots:
            ref = slots[args.pos]
            print(f'[fill-stats] 参战位: {row_label(ref)} = {args.name}' + (' [dry-run]' if args.dry_run else ''))
            if not args.dry_run:
                ws[ref] = args.name
        else:
            print(f'[fill-stats] --pos 无效「{args.pos}」，可选: {"、".join(slots)}（跳过参战位）', file=sys.stderr)
    if args.dry_run:
        wb.close()
        print('[fill-stats] dry-run：未写入任何内容')
        return 0
    wb.save(args.file)
    wb.close()
    print(f'[fill-stats] 已保存: {args.file}')
    return 0


# ============ skills（角色卡技能/宝具一览） ============
def cmd_skills(args):
    if not os.path.exists(args.card):
        print(f'[错误] 角色卡不存在: {args.card}', file=sys.stderr)
        return 1
    wb = openpyxl.load_workbook(args.card, read_only=True, data_only=True)
    sheet = None
    for cand in ('角色卡 ', '角色卡'):
        if cand in wb.sheetnames:
            sheet = wb[cand]
            break
    if sheet is None:
        print(f'[错误] 无「角色卡」sheet', file=sys.stderr)
        wb.close()
        return 1
    print(f'== 角色卡技能/宝具: {os.path.basename(args.card)} ==')

    def field(row, col):
        v = sheet[f'{col}{row}'].value
        return '' if v is None else str(v).strip()

    blocks = []
    max_row = min(sheet.max_row, 140)
    for r in range(1, max_row + 1):
        # 左侧技能块：B='模板'，名称在 C
        if field(r, 'B') == '模板':
            name = field(r, 'C')
            if not name:
                continue
            level = field(r, 'F')  # 等级
            timing = field(r + 2, 'F') if field(r + 2, 'E') == '发动时机' else ''
            mana = field(r + 2, 'I') if field(r + 2, 'H') == '魔力消耗' else ''
            cooldown = field(r + 1, 'I') if field(r + 1, 'H') == '回转' else ''
            stype = field(r + 1, 'F') if field(r + 1, 'E') == '类型' else ''
            effect = field(r + 3, 'C')
            blocks.append(('技能', name, level, stype, timing, mana, cooldown, effect))
        # 右侧宝具/礼装块：L='模板'，名称在 M
        if field(r, 'L') == '模板':
            name = field(r, 'M')
            if not name:
                continue
            level = field(r, 'P')
            timing = field(r + 2, 'P') if field(r + 2, 'O') == '发动时机' else ''
            mana = field(r + 2, 'S') if field(r + 2, 'R') == '魔力消耗' else ''
            cooldown = field(r + 1, 'S') if field(r + 1, 'R') == '回转' else ''
            stype = field(r + 1, 'P') if field(r + 1, 'O') == '类型' else ''
            effect = field(r + 3, 'M')
            blocks.append(('宝具', name, level, stype, timing, mana, cooldown, effect))

    seen = set()
    for kind, name, level, stype, timing, mana, cooldown, effect in blocks:
        key = (kind, name)
        if key in seen:
            continue
        seen.add(key)
        eff = effect[:60] + ('…' if len(effect) > 60 else '')
        print(f'  [{kind}] {name}（等级{level or "?"} 类型{stype or "?"}）')
        print(f'      发动时机: {timing or "?"} | 魔力消耗: {mana or "0"} | 回转: {cooldown or "0"}')
        if eff:
            print(f'      效果: {eff}')
    wb.close()
    return 0


# ============ verdict（追加判定记录） ============
def cmd_verdict(args):
    path = args.file
    os.makedirs(os.path.dirname(os.path.abspath(path)) or '.', exist_ok=True)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    entry = f'- [{now}] {args.append.strip()}'
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
    else:
        content = '# 战斗判定记录\n\n（GM 手动裁决与备注，按时间追加）\n'
    with open(path, 'a', encoding='utf-8') as f:
        f.write('\n' + entry)
    print(f'[verdict] 已追加到 {path}')
    print(entry)
    return 0


# ============ handoff ============
def cmd_handoff(args):
    wb, ws = load_ws(args.file, SHEET_BATTLE)
    title = args.title or os.path.splitext(os.path.basename(args.file))[0]
    out = args.out or os.path.join(os.path.dirname(os.path.abspath(args.file)), f'{title}交接.md')

    blue = [cell(ws, BLUE_SLOTS[k]) for k in ('主力', '辅助1', '辅助2', '辅助3', '辅助4')]
    yellow = [cell(ws, YELLOW_SLOTS[k]) for k in ('主力', '辅助1', '辅助2', '辅助3', '辅助4')]

    lines = []
    lines.append(f'# {title}战斗交接')
    lines.append('')
    lines.append(f'> 更新时间：{datetime.datetime.now():%Y-%m-%d %H:%M}')
    lines.append(f'> 战斗表：`{os.path.abspath(args.file)}`')
    lines.append('')
    lines.append('## 编队')
    lines.append(f'- 蓝方：主力={blue[0]}，辅助={[b for b in blue[1:] if b and b != "无"] or "无"}')
    lines.append(f'- 黄方：主力={yellow[0]}，辅助={[y for y in yellow[1:] if y and y != "无"] or "无"}')
    lines.append('')
    lines.append('## 属性与总值')
    lines.append('| 属性 | 蓝总值 | 黄总值 |')
    lines.append('|---|---|---|')
    for stat, r in STAT_ROWS.items():
        lines.append(f'| {stat} | {cell(ws, f"G{r}")} | {cell(ws, f"M{r}")} |')
    lines.append('')
    lines.append('## 战前修正')
    lines.append(f'- 胜补：蓝={cell(ws, "G14")} 黄={cell(ws, "M14")}')
    lines.append(f'- 胜惩：蓝={cell(ws, "G15")} 黄={cell(ws, "M15")}')
    lines.append(f'- 保底={cell(ws, "R16")} 保穿={cell(ws, "R17")}')
    lines.append('')
    lines.append('## 魔力')
    lines.append(f'- 蓝主力={cell(ws, "R13")} 蓝辅助={[cell(ws, f"{c}13") for c in ("S","T","U","V")]} | 黄主力={cell(ws, "W13")} 黄辅助={[cell(ws, f"{c}13") for c in ("X","Y","Z","AA")]}')
    lines.append(f'- 属性扣除={cell(ws, "R14")} 单独扣除={cell(ws, "R15")}')
    lines.append('')
    lines.append('## 最终工序胜率')
    lines.append('| 列 | 蓝 | 黄 |')
    lines.append('|---|---|---|')
    headers = {c: cell(ws, f'{c}85') for c in 'BCDEFGHIJKLMN'}
    for c in 'BCDEFGHIJKLMN':
        if headers[c]:
            lines.append(f'| {headers[c]} | {cell(ws, f"{c}86")} | {cell(ws, f"{c}87")} |')
    lines.append('')
    if args.verdict and os.path.exists(args.verdict):
        with open(args.verdict, 'r', encoding='utf-8') as f:
            lines.append('## 判定记录')
            lines.append('')
            lines.append(f.read().strip())
            lines.append('')
    if args.note:
        lines.append('## 备注')
        lines.append('')
        lines.append(args.note)
        lines.append('')
    lines.append('## 下次继续')
    lines.append('- 用 battle-sheet-xlsx 技能：`status` 看当前状态、`fill-stats` 从角色卡补属性、`verdict` 记判定、`handoff` 刷新本文件。')
    lines.append('- 核对要点：技能/宝具发动时机顺序（常驻→战斗开始时→初始工序→主要工序→最终工序）、魔力不足惩罚档位、属性优劣→基础胜率→最终胜率链。')
    lines.append('')
    lines.append('## 相关文件')
    lines.append(f'- 战斗表：`{os.path.abspath(args.file)}`')
    if args.verdict:
        lines.append(f'- 判定记录：`{os.path.abspath(args.verdict)}`')
    lines.append('- 技能：`.dsh/skills/battle-sheet-xlsx/` 与 `.claude/skills/battle-sheet-xlsx/`')
    lines.append('')

    with open(out, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    wb.close()
    print(f'[handoff] 已生成战斗交接: {out}')
    return 0


def main():
    parser = argparse.ArgumentParser(description='战斗表 Excel 辅助脚本')
    sub = parser.add_subparsers(dest='command', required=True)

    p_status = sub.add_parser('status')
    p_status.add_argument('--file', required=True)
    p_status.add_argument('--sheet', default=SHEET_BATTLE)

    p_roster = sub.add_parser('roster')
    p_roster.add_argument('--file', required=True)

    p_fill = sub.add_parser('fill-stats')
    p_fill.add_argument('--file', required=True)
    p_fill.add_argument('--card', required=True, help='角色卡 xlsx')
    p_fill.add_argument('--name', required=True, help='名单中的角色名（与战斗表参战位一致）')
    p_fill.add_argument('--side', default='蓝', choices=['蓝', '黄'])
    p_fill.add_argument('--pos', default=None, help='主力/辅助1~辅助4（缺省只填名单不填参战位）')
    p_fill.add_argument('--no-slot', action='store_true', help='只更新名单，不写参战位')
    p_fill.add_argument('--dry-run', action='store_true')

    p_skills = sub.add_parser('skills')
    p_skills.add_argument('--card', required=True)

    p_verdict = sub.add_parser('verdict')
    p_verdict.add_argument('--file', required=True, help='判定记录 md 路径')
    p_verdict.add_argument('--append', required=True, help='追加内容')

    p_handoff = sub.add_parser('handoff')
    p_handoff.add_argument('--file', required=True)
    p_handoff.add_argument('--out', default=None)
    p_handoff.add_argument('--title', default=None)
    p_handoff.add_argument('--note', default=None)
    p_handoff.add_argument('--verdict', default=None, help='判定记录 md，合并进交接文档')

    args = parser.parse_args()
    if args.command == 'status':
        return cmd_status(args)
    if args.command == 'roster':
        return cmd_roster(args)
    if args.command == 'fill-stats':
        return cmd_fill_stats(args)
    if args.command == 'skills':
        return cmd_skills(args)
    if args.command == 'verdict':
        return cmd_verdict(args)
    if args.command == 'handoff':
        return cmd_handoff(args)
    return 1


if __name__ == '__main__':
    sys.exit(main())
